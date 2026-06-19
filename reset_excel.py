"""
Remove linhas de contingência do Excel para permitir nova coleta.
Mantém apenas linhas com origem "CME via TradingView".
"""
import os
from openpyxl import load_workbook

EXCEL_PATH = "data/CME_Futures_Database.xlsx"

if not os.path.exists(EXCEL_PATH):
    print("Excel não existe ainda — nada a limpar.")
else:
    wb = load_workbook(EXCEL_PATH)
    for sheet in wb.sheetnames:
        if sheet == "__tmp__":
            continue
        ws = wb[sheet]
        rows_to_delete = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            source = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            if "CONTINGÊNCIA" in source:
                rows_to_delete.append(i)
        # Deleta de baixo para cima
        for i in reversed(rows_to_delete):
            ws.delete_rows(i)
        print(f"  {sheet}: {len(rows_to_delete)} linhas de contingência removidas")
    wb.save(EXCEL_PATH)
    print("Excel limpo.")
