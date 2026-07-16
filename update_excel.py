"""
Excel Updater
Salva os dados coletados no Excel, com uma aba por commodity.
Inclui coluna "Origem" para distinguir dados reais do CME vs. contingência.
Linhas de contingência são destacadas em amarelo (padrão metodologia GT CVU).
"""

import os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "data/CME_Futures_Database.xlsx"

TAB_COLORS = {
    "HH":       "1F4E79",
    "Brent":    "833C00",
    "NBP":      "375623",
    "JKM":      "7030A0",
    "TTF":      "C55A11",
    "Coal_API2":"44546A",
}

HEADER_BG      = "2E4057"
HEADER_FG      = "FFFFFF"
ALT_ROW        = "EEF2F7"
FALLBACK_ROW   = "FFF2CC"   # Amarelo suave — destaque para contingência
FALLBACK_FONT  = "7F6000"   # Texto dourado escuro nas linhas de contingência
BORDER_CLR     = "B8C4CE"

COLUMNS    = ["Data de Coleta", "Contrato", "Data de Vencimento",
              "Preço de Fechamento (USD)", "Origem"]
COL_WIDTHS = [18, 14, 22, 28, 32]


def _border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, tab_color: str):
    for i, (h, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = tab_color


def _style_row(ws, row_idx: int, is_fallback: bool):
    """
    Estiliza linha de dados.
    Linhas de contingência ficam em amarelo para fácil identificação.
    """
    if is_fallback:
        fill_color = FALLBACK_ROW
        font_color = FALLBACK_FONT
        bold = True
    else:
        fill_color = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
        font_color = "000000"
        bold = False

    for col in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=row_idx, column=col)
        c.fill = PatternFill("solid", start_color=fill_color)
        c.font = Font(name="Arial", size=10, color=font_color, bold=bold)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()


def _get_or_create_sheet(wb: Workbook, ticker: str):
    if ticker in wb.sheetnames:
        return wb[ticker]
    ws = wb.create_sheet(title=ticker)
    _apply_header(ws, TAB_COLORS.get(ticker, "4472C4"))
    return ws


def _existing_keys(ws) -> set:
    """Retorna pares (data_coleta, contrato) já presentes — evita duplicatas."""
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            keys.add((str(row[0]), str(row[1])))
    return keys


def _month_key(contract: str):
    """'JUL 26' -> (2026, 7) para ordenar a curva por vencimento."""
    codes = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
             "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
    try:
        mon, yy = str(contract).strip().split()
        return (2000 + int(yy), codes.get(mon.upper(), 99))
    except Exception:
        return (9999, 99)


def _reorder_sheet(ws, tab_color: str):
    """
    Reordena a aba: data de coleta DECRESCENTE (mais recente no topo);
    dentro de cada coleta, a curva em ordem CRESCENTE de vencimento.
    Reescreve os dados e reaplica o estilo.
    """
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            data.append(list(row[:5]))
    if not data:
        return

    # coleta desc, vencimento (contrato) asc
    data.sort(key=lambda r: (str(r[0]), ), reverse=True)          # 1º: coleta desc
    data.sort(key=lambda r: _month_key(r[1]))                     # 2º: contrato asc
    data.sort(key=lambda r: str(r[0]), reverse=True)             # estável: coleta desc domina

    # limpa linhas de dados
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # reescreve ordenado
    for i, r in enumerate(data, start=2):
        source = str(r[4]) if len(r) > 4 and r[4] is not None else ""
        is_fallback = "CONTINGÊNCIA" in source
        for col in range(5):
            ws.cell(i, col + 1, r[col] if col < len(r) else "")
        _style_row(ws, i, is_fallback)


def update_excel(all_data: dict):
    os.makedirs("data", exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        default = wb.active
        default.title = "__tmp__"
        print(f"Criando novo arquivo: {EXCEL_PATH}")

    total_added = 0
    total_fallback = 0

    for ticker, rows in all_data.items():
        if not rows:
            print(f"  [SKIP] {ticker}: sem dados para salvar")
            continue

        ws = _get_or_create_sheet(wb, ticker)
        existing = _existing_keys(ws)
        added = 0
        fallback = 0

        for row_data in rows:
            key = (row_data["collection_date"], row_data["contract"])
            if key in existing:
                continue

            source = row_data.get("source", "CME via TradingView")
            is_fallback = "CONTINGÊNCIA" in source

            nr = ws.max_row + 1
            ws.cell(nr, 1, row_data["collection_date"])
            ws.cell(nr, 2, row_data["contract"])
            ws.cell(nr, 3, row_data["expiry_date"])
            ws.cell(nr, 4, row_data["settlement_price"])
            ws.cell(nr, 5, source)
            _style_row(ws, nr, is_fallback)

            existing.add(key)
            added += 1
            if is_fallback:
                fallback += 1

        total_added += added
        total_fallback += fallback

        fallback_info = f" ({fallback} contingência)" if fallback else ""
        print(f"  [Excel] {ticker}: {added} linhas adicionadas{fallback_info}")

    # Remove aba temporária se outras foram criadas
    if "__tmp__" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["__tmp__"]

    # Reordena todas as abas de futuros: coleta mais recente no topo,
    # curva em ordem de vencimento dentro de cada dia.
    for ticker in all_data.keys():
        if ticker in wb.sheetnames:
            _reorder_sheet(wb[ticker], TAB_COLORS.get(ticker, "4472C4"))

    if not wb.sheetnames:
        ws = wb.create_sheet("Info")
        ws["A1"] = "Nenhum dado coletado em " + date.today().isoformat()

    wb.save(EXCEL_PATH)
    print(f"\nArquivo salvo: {EXCEL_PATH}")
    print(f"Total de novas linhas: {total_added} ({total_fallback} por contingência)")


if __name__ == "__main__":
    # Teste com dados mistos: CME real + contingência
    today = date.today().isoformat()
    sample = {
        "TTF": [
            {
                "collection_date": today,
                "ticker": "TTF",
                "contract": "JUL 26",
                "expiry_date": "2026-06-30",
                "settlement_price": "11.20",
                "source": "CME",
            },
            {
                "collection_date": today,
                "ticker": "TTF",
                "contract": "DEC 27",
                "expiry_date": "2027-11-30",
                "settlement_price": "10.8500",
                "source": "CONTINGÊNCIA (média 12 meses)",
            },
        ],
        "HH": [
            {
                "collection_date": today,
                "ticker": "HH",
                "contract": "JUL 26",
                "expiry_date": "2026-07-28",
                "settlement_price": "3.245",
                "source": "CME",
            },
        ],
    }
    update_excel(sample)
