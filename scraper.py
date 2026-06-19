"""
CME Futures Scraper — TradingView via Playwright
Versão com logging detalhado para diagnóstico.
"""

import json, os, re, time
from datetime import date
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

EXCEL_PATH = "data/CME_Futures_Database.xlsx"

TICKERS = {
    "HH": {
        "name": "Henry Hub Natural Gas",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-NG1!/contracts/",
        "root": "NG",
        "use_fallback": False,
    },
    "Brent": {
        "name": "Brent Crude Oil",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-BB1!/contracts/",
        "root": "BB",
        "use_fallback": False,
    },
    "NBP": {
        "name": "UK NBP Natural Gas Calendar Month",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-UKG1!/contracts/",
        "root": "UKG",
        "use_fallback": False,
    },
    "JKM": {
        "name": "LNG Japan Korea Marker (Platts)",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-LNG1!/contracts/",
        "root": "LNG",
        "use_fallback": False,
    },
    "TTF": {
        "name": "Dutch TTF Natural Gas (Platts ENDEX)",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-TTF1!/contracts/",
        "root": "TTF",
        "use_fallback": True,
    },
    "Coal_API2": {
        "name": "Coal API2 CIF ARA (Argus/McCloskey)",
        "tv_url": "https://www.tradingview.com/symbols/NYMEX-MTF1!/contracts/",
        "root": "MTF",
        "use_fallback": False,
    },
}

MONTH_CODES = {
    "F":"JAN","G":"FEB","H":"MAR","J":"APR","K":"MAY","M":"JUN",
    "N":"JUL","Q":"AUG","U":"SEP","V":"OCT","X":"NOV","Z":"DEC",
}

INVALID = {"-","","N/A","n/a","—","–","0","0.00","unch","UNCH","null","None"}


# ---------------------------------------------------------------------------
# Fallback: média 12 meses
# ---------------------------------------------------------------------------

def _load_historical(ticker):
    if not os.path.exists(EXCEL_PATH):
        return {}
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception:
        return {}
    if ticker not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[ticker]
    h = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1] or not row[3]:
            continue
        # Ignora linhas de contingência anteriores
        source = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if "CONTINGÊNCIA" in source:
            continue
        try:
            p = float(str(row[3]).replace(",", "."))
        except ValueError:
            continue
        h.setdefault(str(row[1]).strip(), []).append((str(row[0]), p))
    wb.close()
    return {c: [p for _, p in sorted(v)] for c, v in h.items()}


def _fallback(ticker, contract):
    prices = _load_historical(ticker).get(contract, [])
    if not prices:
        return "", ""
    last12 = prices[-12:]
    return f"{sum(last12)/len(last12):.4f}", f"CONTINGÊNCIA (média {len(last12)} meses)"


# ---------------------------------------------------------------------------
# Converte símbolo CME para label legível
# ---------------------------------------------------------------------------

def _symbol_to_label(sym):
    s = sym.split(":")[-1]
    m = re.search(r"[A-Z]+([FGHJKMNQUVXZ])(\d{4})$", s)
    if not m:
        return ""
    month = MONTH_CODES.get(m.group(1), "")
    return f"{month} {m.group(2)[2:]}" if month else ""


# ---------------------------------------------------------------------------
# Coleta via Playwright
# ---------------------------------------------------------------------------

def _scrape_ticker(ticker, config, ctx):
    today = date.today().isoformat()
    results = []
    api_items = []
    all_json_urls = []

    page = ctx.new_page()

    def on_response(resp):
        try:
            ct = resp.headers.get("content-type", "")
            if "json" not in ct or resp.status != 200:
                return
            all_json_urls.append(resp.url)
            body = resp.json()
            # Endpoint scanner do TradingView: {"data":[{"s":"NYMEX:NGN2026","d":[price,...]}]}
            if isinstance(body, dict) and "data" in body:
                items = body["data"]
                if items and isinstance(items[0], dict) and "s" in items[0]:
                    print(f"    [API capturada] {resp.url[:100]} — {len(items)} itens")
                    # Log dos primeiros 3 itens para diagnóstico
                    for item in items[:3]:
                        print(f"      Exemplo: s={item.get('s')} d={item.get('d')}")
                    api_items.extend(items)
        except Exception:
            pass

    page.on("response", on_response)

    try:
        page.goto(config["tv_url"], wait_until="networkidle", timeout=50000)
        page.wait_for_timeout(5000)
    except Exception as e:
        print(f"    [AVISO] goto: {e}")

    page.remove_listener("response", on_response)

    print(f"    JSON URLs capturadas: {len(all_json_urls)}")
    print(f"    Itens da API scanner: {len(api_items)}")

    # --- Processa itens da API scanner ---
    if api_items:
        ok = fb = skip = 0
        for item in api_items:
            sym  = item.get("s", "")
            vals = item.get("d", [])
            label = _symbol_to_label(sym)
            if not label:
                continue

            # d[0] = close/last price, d[1] = expiration_date
            price_raw = vals[0] if vals else None
            expiry    = str(vals[1]) if len(vals) > 1 else ""
            price_str = str(price_raw).strip() if price_raw is not None else ""

            if price_str and price_str not in INVALID and price_str != "None":
                results.append({"collection_date": today, "ticker": ticker,
                    "contract": label, "expiry_date": expiry,
                    "settlement_price": price_str, "source": "CME via TradingView"})
                ok += 1
            elif config["use_fallback"]:
                fp, nota = _fallback(ticker, label)
                if fp:
                    results.append({"collection_date": today, "ticker": ticker,
                        "contract": label, "expiry_date": expiry,
                        "settlement_price": fp, "source": nota})
                    fb += 1
                else:
                    skip += 1
            else:
                skip += 1

        print(f"    Resultado: {ok} com preço, {fb} fallback, {skip} sem preço/histórico")
        page.close()
        return results

    # --- Fallback: lê tabela HTML ---
    print("    API scanner não capturada — tentando HTML da tabela...")
    try:
        # Aguarda mais um pouco e tenta rolar a página
        page.wait_for_timeout(3000)
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(2000)

        # Procura por qualquer elemento que contenha preços
        html = page.content()
        root = config["root"]

        # Busca símbolos no HTML (ex: NGN2026, BBN2026)
        pattern = rf'{root}([FGHJKMNQUVXZ])(\d{{4}})'
        sym_matches = re.findall(pattern, html)
        print(f"    Símbolos {root}* no HTML: {len(sym_matches)} encontrados")
        print(f"    Exemplos: {sym_matches[:5]}")

        # Busca preços próximos dos símbolos
        # Estratégia: extrai blocos de texto que contenham símbolo + número
        blocks = re.findall(
            rf'{root}[FGHJKMNQUVXZ]\d{{4}}[^<]{{0,200}}?(\d+\.?\d*)',
            html
        )
        print(f"    Blocos com preço encontrados: {blocks[:5]}")

        # Tenta ler tabela estruturada
        rows = page.query_selector_all("table tbody tr")
        print(f"    <tr> na tabela: {len(rows)}")

        if rows:
            ok = 0
            for row in rows:
                cells = row.query_selector_all("td")
                if len(cells) < 2:
                    continue
                texts = [c.inner_text().strip() for c in cells]
                print(f"    Row: {texts[:5]}")  # log das primeiras colunas

                sym_text = texts[0] if texts else ""
                label = _symbol_to_label(sym_text) or sym_text

                # Procura preço nas células
                price_str = ""
                for t in texts[1:]:
                    if re.match(r'^\d+\.?\d+$', t) and t not in INVALID:
                        price_str = t
                        break

                if label and price_str:
                    results.append({"collection_date": today, "ticker": ticker,
                        "contract": label, "expiry_date": "",
                        "settlement_price": price_str, "source": "CME via TradingView"})
                    ok += 1

            print(f"    HTML tabela: {ok} contratos extraídos")

    except Exception as e:
        print(f"    [ERRO HTML] {e}")

    page.close()
    return results


def collect_all():
    print(f"\n=== Coleta CME Futures — {date.today().isoformat()} ===\n")
    all_data = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage",
                  "--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900},
            locale="en-US",
            extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
        )

        # Aquece a sessão na home do TradingView
        home = ctx.new_page()
        try:
            home.goto("https://www.tradingview.com/", wait_until="domcontentloaded", timeout=20000)
            home.wait_for_timeout(2000)
            print("Home TradingView: OK\n")
        except Exception as e:
            print(f"Home TradingView: {e}\n")
        home.close()

        for ticker, config in TICKERS.items():
            print(f"Buscando {ticker} ({config['name']})...")
            try:
                rows = _scrape_ticker(ticker, config, ctx)
                all_data[ticker] = rows
                print(f"  [OK] {ticker}: {len(rows)} contratos\n")
            except Exception as e:
                print(f"  [ERRO] {ticker}: {e}\n")
                all_data[ticker] = []
            time.sleep(3)

        browser.close()

    total = sum(len(v) for v in all_data.values())
    print(f"Total de contratos coletados: {total}")
    return all_data


if __name__ == "__main__":
    from update_excel import update_excel
    data = collect_all()
    update_excel(data)
