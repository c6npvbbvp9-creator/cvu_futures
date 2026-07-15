"""
Cotações Diárias — aba resumo com 1 linha por dia e 1 coluna por ticker.
Datas mais recentes aparecem primeiro.

Estrutura da aba "Cotações Diárias":
  Data | HH | HH (contrato) | Brent | Brent (contrato) | ...

CORREÇÕES (vs. versão anterior):
  1. Ignora contratos JÁ VENCIDOS (expiry < data de coleta). O bug original
     capturava o JUL 26 do JKM (vencido em 17/06) com preço congelado em
     18.783, inflando a média mensal.
  2. Trava o contrato de referência DENTRO do mês (REFERENCE_MODE). Evita
     rolagem no meio do mês, que produzia saltos artificiais de ~15%.
  3. Registra o contrato usado em cada dia, para auditoria.

Uso:
  python daily_quotes.py --backfill   → popula histórico via Yahoo Finance
  python daily_quotes.py              → adiciona só o dia de hoje
  python daily_quotes.py --rebuild    → reconstrói a aba inteira a partir
                                        das abas de curva já no Excel
                                        (use isto para corrigir junho)
"""

import sys, time, json, urllib.request
from datetime import date, datetime, timezone
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "data/CME_Futures_Database.xlsx"
SHEET_NAME = "Cotações Diárias"

# Tickers da aba "Cotações Diárias".
# Diesel  → EIA (fonte externa, ver eia_diesel.py) — NÃO vem das abas de curva
# Coal_API2 → curva do CME/TradingView (front month), proxy de CSARM01
TICKERS = ["HH", "Brent", "NBP", "JKM", "TTF", "Coal_API2", "Diesel"]

# Tickers cujo preço vem das abas de curva do Excel.
# Os demais são populados por módulos externos (ex.: Diesel via EIA).
CURVE_TICKERS = ["HH", "Brent", "NBP", "JKM", "TTF", "Coal_API2"]

# ---------------------------------------------------------------------------
# Metodologia do contrato de referência
# ---------------------------------------------------------------------------
#   "front_month" → contrato ativo mais próximo do vencimento (rola quando vence)
#   "m_plus_1"    → contrato do mês SEGUINTE ao mês de coleta; fica travado o
#                   mês inteiro, rolando só na virada do mês. É o que produz
#                   uma média mensal comparável ao settlement oficial.
#
# Confirme com a gestora qual metodologia gera o 17.16 antes de fechar o número.
REFERENCE_MODE = "m_plus_1"

YAHOO_SYMBOLS = {
    "HH":       "NG=F",
    "Brent":    "BZ=F",
    "NBP":      "UKG=F",
    "JKM":      "JKM=F",
    "TTF":      "TTF=F",
}
# Coal_API2 NÃO usa Yahoo: vem da curva TradingView (ICEEUR-ATW1!), como os
# demais futuros. O MTF=F do Yahoo tem histórico congelado e foi descartado.
# NOTA: o backfill via Yahoo usa contrato CONTÍNUO (rolagem automática do
# provedor) e preço de fechamento, não settlement. Foi a origem do erro de
# junho no JKM. Prefira o rebuild a partir das abas de curva (--rebuild).

MONTH_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4,  "MAY": 5,  "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

HEADER_BG  = "2E4057"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EEF2F7"
BORDER_CLR = "B8C4CE"
CONTRACT_FG = "6B7A88"   # cinza para as colunas de contrato (metadado)


# ---------------------------------------------------------------------------
# Formatação Excel
# ---------------------------------------------------------------------------

def _columns():
    """Data | <ticker> | <ticker> (contrato) | ..."""
    cols = ["Data"]
    for t in TICKERS:
        cols += [t, f"{t} (contrato)"]
    return cols


def _border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _setup_sheet(ws):
    headers = _columns()
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        c.fill      = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[get_column_letter(i)].width = 14 if i == 1 else (
            13 if "(contrato)" not in h else 15)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"
    ws.sheet_properties.tabColor = "2E86AB"


def _style_row(ws, row_idx):
    fill = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
    ncols = len(_columns())
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.fill      = PatternFill("solid", start_color=fill)
        # colunas pares (a partir da 3) são "(contrato)" → cinza, itálico
        is_contract = col >= 3 and col % 2 == 1
        c.font      = Font(name="Arial", size=10,
                           color=CONTRACT_FG if is_contract else "000000",
                           italic=is_contract)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()


# ---------------------------------------------------------------------------
# Seleção do contrato de referência
# ---------------------------------------------------------------------------

def _contract_month(label: str):
    """'JUL 26' -> (2026, 7). Retorna None se não parsear."""
    try:
        mon, yy = label.strip().split()
        return (2000 + int(yy), MONTH_NUM[mon.upper()])
    except Exception:
        return None


def _is_expired(expiry: str, ref_date: str) -> bool:
    """True se o contrato já venceu na data de referência."""
    e = str(expiry or "").strip()[:10]
    if len(e) != 10:
        return False          # sem expiry conhecido → não descarta
    return e < ref_date


def _pick_reference(rows: list, ref_date: str):
    """
    rows: lista de tuplas da aba de curva
          (data_coleta, contrato, vencimento, preço, origem)
    Retorna (preço_str, contrato_str) ou ("", "").
    """
    # 1. descarta contratos já vencidos e linhas sem preço
    live = [r for r in rows
            if r[3] not in (None, "")
            and not _is_expired(r[2], ref_date)]
    if not live:
        return "", ""

    # 2. ordena por mês do contrato (não pela ordem da tabela, que é instável)
    def sort_key(r):
        cm = _contract_month(str(r[1]))
        return cm if cm else (9999, 99)

    live.sort(key=sort_key)

    if REFERENCE_MODE == "front_month":
        chosen = live[0]
    else:  # m_plus_1 — trava no mês seguinte ao mês de coleta
        y, m = int(ref_date[:4]), int(ref_date[5:7])
        target = (y + 1, 1) if m == 12 else (y, m + 1)
        chosen = next((r for r in live if _contract_month(str(r[1])) == target),
                      live[0])   # fallback: front month se o alvo não existir

    return str(chosen[3]), str(chosen[1])


def _prices_for_date(wb, ref_date: str) -> dict:
    """
    {ticker: (preço, contrato)} lendo as abas de curva do Excel.
    Tickers fora de CURVE_TICKERS (ex.: Diesel/EIA) ficam vazios aqui —
    são preenchidos pelos seus próprios módulos, sem serem sobrescritos.
    """
    out = {t: ("", "") for t in TICKERS}
    for ticker in CURVE_TICKERS:
        if ticker not in wb.sheetnames:
            continue
        rows = [r for r in wb[ticker].iter_rows(min_row=2, values_only=True)
                if r[0] and str(r[0])[:10] == ref_date]
        if rows:
            out[ticker] = _pick_reference(rows, ref_date)
    return out


def _get_today_from_excel() -> dict:
    """Contrato de referência de hoje, por ticker."""
    today = date.today().isoformat()
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        prices = _prices_for_date(wb, today)
        wb.close()
        return prices
    except Exception as e:
        print(f"  [AVISO] {e}")
        return {t: ("", "") for t in TICKERS}


# ---------------------------------------------------------------------------
# Atualização do Excel
# ---------------------------------------------------------------------------

def _rebuild_sheet(wb, by_date: dict):
    """
    by_date: {date_str: {ticker: (preço, contrato)}}
    Reconstrói a aba, mais recente primeiro.
    """
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws = wb.create_sheet(title=SHEET_NAME, index=0)
    _setup_sheet(ws)

    for row_idx, d in enumerate(sorted(by_date.keys(), reverse=True), start=2):
        vals = by_date[d]
        ws.cell(row_idx, 1, d)
        col = 2
        for ticker in TICKERS:
            price, contract = vals.get(ticker, ("", ""))
            ws.cell(row_idx, col,     price)
            ws.cell(row_idx, col + 1, contract)
            col += 2
        _style_row(ws, row_idx)

    return ws


def _read_existing(wb) -> dict:
    """Lê a aba atual no formato {date: {ticker: (preço, contrato)}}."""
    existing = {}
    if SHEET_NAME not in wb.sheetnames:
        return existing

    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    has_contract_cols = any("(contrato)" in str(h) for h in header)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = str(row[0])[:10]
        vals = {}
        for i, ticker in enumerate(TICKERS):
            if has_contract_cols:
                pi, ci = 1 + i * 2, 2 + i * 2
                price    = row[pi] if len(row) > pi else ""
                contract = row[ci] if len(row) > ci else ""
            else:  # formato antigo: só preço
                pi = 1 + i
                price, contract = (row[pi] if len(row) > pi else ""), ""
            vals[ticker] = (
                str(price) if price not in (None, "") else "",
                str(contract) if contract not in (None, "") else "",
            )
        existing[d] = vals
    return existing


def update_daily_quotes(new_prices: dict, overwrite: bool = False):
    """
    new_prices: {date_str: {ticker: (preço, contrato)}}
    overwrite=True substitui valores existentes (use ao corrigir dados errados).
    """
    wb = load_workbook(EXCEL_PATH)
    merged = _read_existing(wb)

    added = updated = 0
    for d, vals in new_prices.items():
        if d not in merged:
            merged[d] = vals
            added += 1
            continue
        for ticker, (price, contract) in vals.items():
            if not price:
                continue
            old_price, _ = merged[d].get(ticker, ("", ""))
            if overwrite or not old_price:
                if old_price != price:
                    updated += 1
                merged[d][ticker] = (price, contract)

    _rebuild_sheet(wb, merged)
    wb.save(EXCEL_PATH)
    print(f"  [Cotações Diárias] {added} dias novos, {updated} valores "
          f"corrigidos ({len(merged)} dias no total)")


# ---------------------------------------------------------------------------
# Coleta via Yahoo (backfill de lacunas)
# ---------------------------------------------------------------------------

def _fetch_yahoo(symbol: str, date_from: str, date_to: str) -> dict:
    """Retorna {date_str: price_str} via API v8 do Yahoo Finance."""
    p1 = int(datetime.strptime(date_from, "%Y-%m-%d")
             .replace(tzinfo=timezone.utc).timestamp())
    p2 = int(datetime.strptime(date_to, "%Y-%m-%d")
             .replace(tzinfo=timezone.utc).timestamp()) + 86400

    for host in ["query1", "query2"]:
        url = (f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}"
               f"?interval=1d&period1={p1}&period2={p2}&includePrePost=false")
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36",
                "Accept": "application/json",
            })
            raw  = urllib.request.urlopen(req, timeout=20).read()
            data = json.loads(raw)
            r    = data["chart"]["result"][0]
            timestamps = r.get("timestamp", [])
            closes     = r["indicators"]["quote"][0].get("close", [])
            result = {}
            for t, c in zip(timestamps, closes):
                if c is None:
                    continue
                d = datetime.fromtimestamp(int(t), tz=timezone.utc).date().isoformat()
                result[d] = f"{float(c):.4f}"
            return result
        except Exception as e:
            print(f"    [{host}] ERRO: {e}")
            time.sleep(1)
    return {}


# ---------------------------------------------------------------------------
# Modos de execução
# ---------------------------------------------------------------------------

def run_today():
    """Adiciona o dia de hoje a partir das abas de curva já no Excel."""
    print("\n[Cotações Diárias] Atualizando com dados de hoje...")
    today  = date.today().isoformat()
    prices = _get_today_from_excel()
    for t in CURVE_TICKERS:
        p, c = prices.get(t, ("", ""))
        print(f"  {t:10s} {p or '—':>10s}  ({c or 'sem contrato'})")
    print("  Diesel     → via eia_diesel.py (fonte externa)")
    update_daily_quotes({today: prices})


def run_rebuild(date_from: str = None, date_to: str = None):
    """
    Reconstrói a aba inteira a partir das abas de curva do Excel,
    aplicando a metodologia corrigida. Sobrescreve valores errados.
    Use isto para consertar junho.
    """
    print("\n[Cotações Diárias] REBUILD a partir das abas de curva "
          f"(modo: {REFERENCE_MODE})...\n")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # todas as datas de coleta presentes em qualquer aba de curva
    all_dates = set()
    for ticker in TICKERS:
        if ticker not in wb.sheetnames:
            continue
        for row in wb[ticker].iter_rows(min_row=2, values_only=True):
            if row[0]:
                all_dates.add(str(row[0])[:10])

    dates = sorted(d for d in all_dates
                   if (not date_from or d >= date_from)
                   and (not date_to or d <= date_to))
    print(f"  {len(dates)} datas de coleta encontradas")

    by_date = {d: _prices_for_date(wb, d) for d in dates}
    wb.close()

    update_daily_quotes(by_date, overwrite=True)

    # Diagnóstico: média por mês e por ticker
    print("\n  Médias mensais recalculadas:")
    months = sorted({d[:7] for d in by_date})
    for mth in months:
        parts = []
        for t in TICKERS:
            vals = [float(by_date[d][t][0])
                    for d in by_date
                    if d[:7] == mth and by_date[d][t][0]]
            parts.append(f"{t}={sum(vals)/len(vals):.2f} (n={len(vals)})"
                         if vals else f"{t}=—")
        print(f"    {mth}: " + "  ".join(parts))


def run_backfill(date_from="2026-06-01", date_to=None):
    """Preenche lacunas via Yahoo Finance (contrato contínuo)."""
    if date_to is None:
        date_to = date.today().isoformat()
    print(f"\n[Cotações Diárias] Backfill {date_from} → {date_to} "
          "via Yahoo Finance...\n")

    all_history = {}
    for ticker, sym in YAHOO_SYMBOLS.items():
        print(f"  {ticker} ({sym})...")
        history = _fetch_yahoo(sym, date_from, date_to)
        filtered = {d: v for d, v in history.items() if date_from <= d <= date_to}
        all_history[ticker] = filtered
        print(f"    {len(filtered)} dias coletados")
        if not filtered:
            print("    [AVISO] símbolo sem dados no Yahoo — pode ser ilíquido")
        time.sleep(1)

    all_dates = set()
    for h in all_history.values():
        all_dates.update(h.keys())

    by_date = {
        d: {t: (all_history.get(t, {}).get(d, ""), "Yahoo contínuo")
            for t in TICKERS}
        for d in sorted(all_dates)
    }

    print(f"\n  {len(by_date)} dias no total")
    if by_date:
        # overwrite=False: não sobrescreve o que veio do CME, só preenche lacunas
        update_daily_quotes(by_date, overwrite=False)
    else:
        print("  Nenhum dado — verifique a conexão com o Yahoo Finance")


if __name__ == "__main__":
    if "--rebuild" in sys.argv:
        run_rebuild()
    elif "--backfill" in sys.argv:
        run_backfill()
    else:
        run_today()
