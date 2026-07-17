"""
CVU Updater — Atualização Diária de Tickers (PREÇOS SPOT)
==========================================================
Fontes SPOT diárias:
  - PTAX   → API BCB (oficial, spot fechamento do dia)
  - HH     → EIA API série RNGWHHD (spot diário oficial)
             Fallback: EnergyRiskIQ CSV
  - Brent  → EIA API série RBRTE (spot diário FOB oficial)
             Fallback: EnergyRiskIQ CSV
  - TTF    → EnergyRiskIQ CSV (spot day-ahead)
             Fallback: Yahoo Finance TTF=F (front-month, menos preciso)
  - NBP    → proxy TTF spot (sem fonte spot gratuita disponível)
  - JKM    → EnergyRiskIQ CSV (spot day-ahead)
  - Óleo/Diesel → GitHub secrets (manual mensal)

Uso:
  python atualizar_cvu.py                    # atualiza ambas colunas
  python atualizar_cvu.py --modo pmo
  python atualizar_cvu.py --modo revisao
  python atualizar_cvu.py --dry-run
"""

import argparse
import csv
import io
import json
import logging
import os
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

import requests
import yfinance as yf
import openpyxl

try:
    from crosscheck_jkm import executar_crosscheck, imprimir_crosscheck
    CROSSCHECK_DISPONIVEL = True
except ImportError:
    CROSSCHECK_DISPONIVEL = False

try:
    from historico_cotacoes import registrar_cotacao
    HISTORICO_DISPONIVEL = True
except ImportError:
    HISTORICO_DISPONIVEL = False

try:
    from database_mensal import atualizar_database_mensal, coletar_dados_mensais
    DB_MENSAL_DISPONIVEL = True
except ImportError:
    DB_MENSAL_DISPONIVEL = False

# ── Configuração ─────────────────────────────────────────────────────────────
EXCEL_PATH = Path(__file__).parent / "CVU_calculo.xlsx"
LOG_PATH   = Path(__file__).parent / "logs" / "cvu_update.log"
CACHE_PATH = Path(__file__).parent / "cache" / "cotacoes.json"

LOG_PATH.parent.mkdir(exist_ok=True)
(Path(__file__).parent / "cache").mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── EIA API — séries spot oficiais ───────────────────────────────────────────
EIA_SERIES = {
    "HH":    "RNGWHHD",   # Henry Hub Natural Gas Spot Price (US$/MMBTU)
    "Brent": "RBRTE",     # Europe Brent Spot Price FOB (US$/bbl)
}

# ── EnergyRiskIQ — CSVs spot gratuitos ──────────────────────────────────────
# EnergyRiskIQ — apenas JKM e TTF têm CSVs públicos
ERIQ_URLS = {
    "JKM": "https://energyriskiq.com/api/jkm-lng-spot-price.csv",
    "TTF": "https://energyriskiq.com/api/ttf-gas-prices.csv",
}

# ── Células — aba DADOS - ATUALIZAR ─────────────────────────────────────────
CELLS_PMO = {
    "PTAX": (21,2), "HH": (22,2), "Brent": (23,2),
    "NBP":  (24,2), "JKM": (25,2), "TTF": (26,2),
    "OleoComb": (27,2), "Diesel": (28,2),
}
CELLS_REVISAO = {
    "PTAX": (21,3), "HH": (22,3), "Brent": (23,3),
    "NBP":  (24,3), "JKM": (25,3), "TTF": (26,3),
    "OleoComb": (27,3), "Diesel": (28,3),
}
SHEET_DADOS = "DADOS - ATUALIZAR"


# ════════════════════════════════════════════════════════════════════════════
# UTILITÁRIO — último dia útil
# ════════════════════════════════════════════════════════════════════════════

def _ultimo_dia_util() -> date:
    d = date.today() - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


# ════════════════════════════════════════════════════════════════════════════
# 1. PTAX — BCB (spot oficial)
# ════════════════════════════════════════════════════════════════════════════

def buscar_ptax() -> float:
    """PTAX de venda — último boletim do dia útil mais recente. API BCB oficial."""
    d = _ultimo_dia_util()
    for _ in range(5):
        url = (
            "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
            f"CotacaoDolarDia(dataCotacao=@dataCotacao)"
            f"?@dataCotacao='{d.strftime('%m-%d-%Y')}'"
            "&$format=json&$select=cotacaoVenda,dataHoraCotacao"
        )
        log.info(f"PTAX [BCB spot]: buscando {d}")
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        registros = r.json().get("value", [])
        if registros:
            valor = round(registros[-1]["cotacaoVenda"], 6)
            log.info(f"PTAX spot ({d}): {valor:.6f} R$/US$")
            return valor
        d -= timedelta(days=1)
        while d.weekday() >= 5:
            d -= timedelta(days=1)
    raise ValueError("PTAX: sem dados nos últimos 5 dias úteis")


# ════════════════════════════════════════════════════════════════════════════
# 2. EIA API — HH e Brent spot oficiais
# ════════════════════════════════════════════════════════════════════════════

def buscar_eia_spot(variavel: str, api_key: str) -> float | None:
    """
    Preço SPOT diário via API oficial do EIA.
    HH:    RNGWHHD — Henry Hub Natural Gas Spot Price (US$/MMBTU)
    Brent: RBRTE   — Europe Brent Spot Price FOB (US$/bbl)
    Chave gratuita: https://www.eia.gov/opendata/register.php
    """
    serie = EIA_SERIES.get(variavel)
    if not serie:
        return None
    log.info(f"{variavel} [EIA spot oficial]: série {serie}")
    url = (
        f"https://api.eia.gov/v2/seriesid/{serie}"
        f"?api_key={api_key}"
        f"&data[]=value"
        f"&sort[0][column]=period"
        f"&sort[0][direction]=desc"
        f"&length=5"
    )
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
        dados = r.json().get("response", {}).get("data", [])
        if dados:
            valor = round(float(dados[0]["value"]), 6)
            periodo = dados[0].get("period", "?")
            unidade = "US$/MMBTU" if variavel == "HH" else "US$/bbl"
            log.info(f"{variavel} spot EIA ({periodo}): {valor:.4f} {unidade}")
            return valor
    except Exception as e:
        log.error(f"{variavel} EIA ERRO: {e}")
    return None


# ════════════════════════════════════════════════════════════════════════════
# 3. EnergyRiskIQ — TTF, JKM, HH e Brent spot (fallback)
# ════════════════════════════════════════════════════════════════════════════

def buscar_eriq_spot(variavel: str,
                     price_col: str = None,
                     conv_eur_mwh: bool = False) -> float | None:
    """
    Preço spot via EnergyRiskIQ CSV (fonte: OilPriceAPI/oilprice.com).
    Filtra heartbeats (fins de semana sem negociação).

    conv_eur_mwh: se True, converte EUR/MWh → US$/MMBTU (para TTF/NBP)
    """
    url = ERIQ_URLS.get(variavel)
    if not url:
        log.warning(f"EnergyRiskIQ: URL não mapeada para {variavel}")
        return None

    log.info(f"{variavel} [EnergyRiskIQ spot]: {url}")
    try:
        r = requests.get(url, timeout=15)
        r.raise_for_status()
    except Exception as e:
        log.error(f"{variavel} EnergyRiskIQ ERRO: {e}")
        return None

    reader = csv.DictReader(io.StringIO(r.text))
    registros = []
    for row in reader:
        try:
            if row.get("Source", "").strip().lower() == "heartbeat":
                continue
            d = datetime.strptime(row["Date"].strip(), "%Y-%m-%d").date()
            # Tenta a coluna de preço correta
            col = price_col or next(
                (k for k in row if "price" in k.lower() or "Price" in k), None
            )
            if not col:
                continue
            preco = float(row[col].replace("$", "").replace("€", "").strip())
            registros.append((d, preco))
        except (ValueError, KeyError, StopIteration):
            continue

    if not registros:
        log.warning(f"{variavel} EnergyRiskIQ: sem dados disponíveis")
        return None

    registros.sort(reverse=True)
    data_ref, valor_raw = registros[0]

    if conv_eur_mwh:
        # Converte EUR/MWh → US$/MMBTU (÷3.41214 × taxa EUR/USD ≈ 1.08)
        eur_usd = 1.08
        valor = round(valor_raw / 3.41214 * eur_usd, 6)
        log.info(f"{variavel} spot EnergyRiskIQ ({data_ref}): {valor_raw:.4f} EUR/MWh → {valor:.4f} US$/MMBTU")
    else:
        valor = round(valor_raw, 6)
        unidade = "US$/MMBTU" if variavel in ("HH", "JKM") else "US$/bbl"
        log.info(f"{variavel} spot EnergyRiskIQ ({data_ref}): {valor:.4f} {unidade}")

    return valor


# ════════════════════════════════════════════════════════════════════════════
# 4. Yahoo Finance — fallback para TTF
# ════════════════════════════════════════════════════════════════════════════

def buscar_yahoo_fallback(variavel: str) -> float | None:
    """
    Fallback via Yahoo Finance — front-month futures (menos preciso que spot).
    Usado apenas quando EIA e EnergyRiskIQ falham.
    """
    tickers = {"TTF": "TTF=F", "HH": "NG=F", "Brent": "BZ=F"}
    ticker_sym = tickers.get(variavel)
    if not ticker_sym:
        return None

    d = _ultimo_dia_util()
    log.warning(f"{variavel} [Yahoo Finance FALLBACK - front-month futuro]: {ticker_sym}")
    try:
        ticker = yf.Ticker(ticker_sym)
        hist = ticker.history(
            start=(d - timedelta(days=7)).isoformat(),
            end=(d + timedelta(days=2)).isoformat()
        )
        if hist.empty:
            return None
        hist.index = hist.index.date
        valor_raw = float(hist.loc[max(hist.index), "Close"])
        if variavel == "TTF":
            valor = round(valor_raw / 3.41214 * 1.08, 6)
            log.warning(f"TTF fallback ({max(hist.index)}): {valor_raw:.4f} EUR/MWh → {valor:.4f} US$/MMBTU [FUTURO, não spot]")
        else:
            valor = round(valor_raw, 6)
            log.warning(f"{variavel} fallback ({max(hist.index)}): {valor:.4f} [FUTURO, não spot]")
        return valor
    except Exception as e:
        log.error(f"{variavel} Yahoo fallback ERRO: {e}")
        return None


# ════════════════════════════════════════════════════════════════════════════
# 5. Funções de coleta por ticker
# ════════════════════════════════════════════════════════════════════════════

def buscar_hh(eia_key: str | None) -> float | None:
    """HH spot: EIA API oficial → Yahoo Finance front-month (fallback)"""
    if eia_key:
        v = buscar_eia_spot("HH", eia_key)
        if v:
            return v
        log.warning("HH: EIA falhou, usando Yahoo Finance (futuro como fallback)")
    return buscar_yahoo_fallback("HH")


def buscar_brent(eia_key: str | None) -> float | None:
    """Brent spot FOB: EIA API oficial → Yahoo Finance front-month (fallback)"""
    if eia_key:
        v = buscar_eia_spot("Brent", eia_key)
        if v:
            return v
        log.warning("Brent: EIA falhou, usando Yahoo Finance (futuro como fallback)")
    return buscar_yahoo_fallback("Brent")


def buscar_ttf() -> float | None:
    """TTF spot day-ahead: EnergyRiskIQ → Yahoo Finance (fallback)"""
    v = buscar_eriq_spot("TTF", price_col="TTF Price (EUR/MWh)", conv_eur_mwh=True)
    if v:
        return v
    log.warning("TTF: EnergyRiskIQ falhou, usando Yahoo Finance (futuro)")
    return buscar_yahoo_fallback("TTF")


def buscar_nbp(ttf_valor: float | None) -> float | None:
    """NBP: proxy do TTF spot (sem fonte spot gratuita disponível)"""
    if ttf_valor:
        log.info(f"NBP [proxy TTF spot]: {ttf_valor:.4f} US$/MMBTU (correlação >95%)")
    return ttf_valor


def buscar_jkm() -> float | None:
    """JKM spot: EnergyRiskIQ CSV"""
    return buscar_eriq_spot("JKM", price_col="JKM Price ($/MMBtu)")


def buscar_anp_manual() -> dict:
    """Óleo e Diesel: lê dos GitHub secrets (manual mensal)"""
    oleo  = os.environ.get("OLEO_COMB")
    diesel = os.environ.get("DIESEL")
    if oleo or diesel:
        log.info("ANP Óleo/Diesel: lendo dos secrets do GitHub.")
        return {
            "OleoComb": round(float(oleo), 4)   if oleo   else None,
            "Diesel":   round(float(diesel), 4) if diesel else None,
        }
    log.warning("ANP Óleo/Diesel: secrets não configurados.")
    return {"OleoComb": None, "Diesel": None}


# ════════════════════════════════════════════════════════════════════════════
# 6. Coleta consolidada
# ════════════════════════════════════════════════════════════════════════════

def coletar_cotacoes(eia_key: str | None = None,
                     jkm_key: str | None = None) -> dict:
    d_ref = _ultimo_dia_util()
    log.info(f"=== Coleta SPOT diária CVU — ref: {d_ref} | EIA: {'✓' if eia_key else '✗'} ===")
    resultado = {}

    # PTAX (BCB spot oficial)
    try:
        resultado["PTAX"] = buscar_ptax()
    except Exception as e:
        log.error(f"PTAX ERRO: {e}"); resultado["PTAX"] = None

    # HH spot (EIA → EnergyRiskIQ → Yahoo)
    try:
        resultado["HH"] = buscar_hh(eia_key)
    except Exception as e:
        log.error(f"HH ERRO: {e}"); resultado["HH"] = None

    # Brent spot (EIA → EnergyRiskIQ → Yahoo)
    try:
        resultado["Brent"] = buscar_brent(eia_key)
    except Exception as e:
        log.error(f"Brent ERRO: {e}"); resultado["Brent"] = None

    # TTF spot (EnergyRiskIQ → Yahoo)
    try:
        resultado["TTF"] = buscar_ttf()
    except Exception as e:
        log.error(f"TTF ERRO: {e}"); resultado["TTF"] = None

    # NBP spot (proxy TTF)
    resultado["NBP"] = buscar_nbp(resultado.get("TTF"))

    # JKM spot (EnergyRiskIQ)
    try:
        resultado["JKM"] = buscar_jkm()
    except Exception as e:
        log.error(f"JKM ERRO: {e}"); resultado["JKM"] = None

    # ANP (manual mensal)
    anp = buscar_anp_manual()
    resultado["OleoComb"] = anp["OleoComb"]
    resultado["Diesel"]   = anp["Diesel"]

    resultado["_atualizado_em"]   = datetime.now().isoformat()
    resultado["_data_referencia"] = d_ref.isoformat()
    resultado["_eia_ativo"]       = bool(eia_key)

    # Cross-check JKM-TTF
    if CROSSCHECK_DISPONIVEL:
        try:
            cc = executar_crosscheck(
                jkm_valor=resultado.get("JKM"),
                ttf_valor=resultado.get("TTF"),
                mes=d_ref.month,
            )
            resultado["_crosscheck_jkm"] = cc
            if cc["status"] != "OK":
                log.warning(f"CROSS-CHECK [{cc['status']}]: {cc['mensagem']}")
        except Exception as e:
            log.warning(f"Cross-check não executado: {e}")
            resultado["_crosscheck_jkm"] = None

    # Histórico diário
    if HISTORICO_DISPONIVEL:
        try:
            registrar_cotacao(resultado)
        except Exception as e:
            log.warning(f"Histórico não atualizado: {e}")

    log.info(f"Coleta concluída: {json.dumps({k:v for k,v in resultado.items() if not k.startswith('_')}, indent=2)}")
    return resultado


# ════════════════════════════════════════════════════════════════════════════
# 7. Excel
# ════════════════════════════════════════════════════════════════════════════

def salvar_cache(cotacoes: dict):
    CACHE_PATH.write_text(json.dumps(cotacoes, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"Cache salvo: {CACHE_PATH}")


def _escrever_celulas(ws, cells: dict, cotacoes: dict):
    for var, (row, col) in cells.items():
        valor = cotacoes.get(var)
        if valor is None:
            continue
        ant = ws.cell(row=row, column=col).value
        ws.cell(row=row, column=col).value = valor
        log.info(f"  {var}: {ant} → {valor} ({ws.cell(row=row,column=col).coordinate})")


def atualizar_excel(cotacoes: dict, modo: str = "ambos"):
    if not EXCEL_PATH.exists():
        log.error(f"Excel não encontrado: {EXCEL_PATH}"); return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    if SHEET_DADOS not in wb.sheetnames:
        log.error(f"Aba '{SHEET_DADOS}' não encontrada."); wb.close(); return

    ws = wb[SHEET_DADOS]
    log.info(f"Atualizando '{SHEET_DADOS}':")
    if modo in ("pmo", "ambos"):
        log.info("  Coluna PMO (B):"); _escrever_celulas(ws, CELLS_PMO, cotacoes)
    if modo in ("revisao", "ambos"):
        log.info("  Coluna Revisão (C):"); _escrever_celulas(ws, CELLS_REVISAO, cotacoes)

    wb.save(EXCEL_PATH)
    wb.close()
    log.info(f"Excel salvo: {EXCEL_PATH}")


# ════════════════════════════════════════════════════════════════════════════
# 8. Relatório
# ════════════════════════════════════════════════════════════════════════════

def imprimir_relatorio(cotacoes: dict):
    eia_ativo = cotacoes.get("_eia_ativo", False)
    d_ref = cotacoes.get("_data_referencia", "?")

    print("\n" + "═"*62)
    print(f"  COTAÇÕES CVU — PREÇOS SPOT — {cotacoes.get('_atualizado_em','')[:16]}")
    print(f"  Data de referência: {d_ref} (último dia útil)")
    print("═"*62)

    fonte_hh_brent = "✓ EIA spot oficial" if eia_ativo else "~ EnergyRiskIQ/Yahoo"
    linhas = [
        ("PTAX (R$/US$)",     "PTAX",     ".6f", "✓ BCB spot oficial"),
        ("HH (US$/MMBTU)",    "HH",       ".4f", fonte_hh_brent),
        ("Brent (US$/bbl)",   "Brent",    ".4f", fonte_hh_brent),
        ("NBP (US$/MMBTU)",   "NBP",      ".4f", "~ proxy TTF spot"),
        ("JKM (US$/MMBTU)",   "JKM",      ".4f", "~ EnergyRiskIQ spot"),
        ("TTF (US$/MMBTU)",   "TTF",      ".4f", "~ EnergyRiskIQ spot"),
        ("Óleo Comb (R$/m³)", "OleoComb", ".2f", "! manual/ANP mensal"),
        ("Diesel (R$/L)",     "Diesel",   ".4f", "! manual/ANP mensal"),
    ]
    for label, key, fmt, origem in linhas:
        v = cotacoes.get(key)
        val_str = f"{v:{fmt}}" if v is not None else "⚠ PENDENTE"
        print(f"  {label:<24} {val_str:<12} {origem}")

    print("═"*62)
    if not eia_ativo:
        print("  ℹ HH/Brent: adicione EIA_API_KEY nos secrets para fonte oficial.")

    cc = cotacoes.get("_crosscheck_jkm")
    if cc and CROSSCHECK_DISPONIVEL:
        imprimir_crosscheck(cc)
    print()


# ════════════════════════════════════════════════════════════════════════════
# 9. DATABASE diária
# ════════════════════════════════════════════════════════════════════════════

SHEET_DATABASE = "DATABASE"
DB_COLUNAS = [
    ("Data",             "DD/MM/YYYY", 14),
    ("PTAX (R$/US$)",    "0.000000",   13),
    ("HH (US$/MMBTU)",   "0.0000",     13),
    ("Brent (US$/bbl)",  "0.0000",     13),
    ("NBP (US$/MMBTU)",  "0.0000",     13),
    ("JKM (US$/MMBTU)",  "0.0000",     13),
    ("TTF (US$/MMBTU)",  "0.0000",     13),
]
FILL_LINHA  = "FFFFFF99"
FILL_HEADER = "1F3864"
FONT_HEADER = "FFFFFF"
FONT_NAME   = "Aptos Narrow"
FONT_SIZE   = 11


def _criar_aba_database(wb):
    from openpyxl.styles import PatternFill, Font, Alignment
    ws = wb.create_sheet(SHEET_DATABASE, 1)
    for col, (nome, fmt, larg) in enumerate(DB_COLUNAS, 1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.fill      = PatternFill("solid", fgColor=FILL_HEADER)
        cell.font      = Font(bold=True, color=FONT_HEADER, name=FONT_NAME, size=FONT_SIZE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    return ws


def _escrever_linha_database(ws, row, cotacoes, d_ref):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime as dt
    try:
        data_val = dt.strptime(str(d_ref)[:10], "%Y-%m-%d").date()
    except Exception:
        data_val = d_ref
    valores = [
        data_val,
        cotacoes.get("PTAX"),
        cotacoes.get("HH"),
        cotacoes.get("Brent"),
        cotacoes.get("NBP"),
        cotacoes.get("JKM"),
        cotacoes.get("TTF"),
    ]
    fill  = PatternFill("solid", fgColor=FILL_LINHA)
    borda = Border(bottom=Side(style="thin", color="CCCCCC"))
    for col, (valor, (nome, fmt, larg)) in enumerate(zip(valores, DB_COLUNAS), 1):
        cell = ws.cell(row=row, column=col, value=valor)
        cell.fill          = fill
        cell.font          = Font(name=FONT_NAME, size=FONT_SIZE)
        cell.alignment     = Alignment(horizontal="center", vertical="center")
        cell.border        = borda
        cell.number_format = fmt


def atualizar_database(cotacoes, excel_path=None):
    path = excel_path or EXCEL_PATH
    if not path.exists():
        log.error(f"Excel não encontrado para DATABASE: {path}"); return

    wb = openpyxl.load_workbook(path)
    if SHEET_DATABASE not in wb.sheetnames:
        log.info(f"Criando aba '{SHEET_DATABASE}'")
        _criar_aba_database(wb)

    ws  = wb[SHEET_DATABASE]
    d_ref = str(cotacoes.get("_data_referencia", date.today().isoformat()))[:10]

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is None:
            continue
        val_str = val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else str(val)[:10]
        if val_str == d_ref:
            log.info(f"DATABASE: atualizando {d_ref} (linha {row})")
            _escrever_linha_database(ws, row, cotacoes, d_ref)
            wb.save(path); wb.close(); return

    ws.insert_rows(2)
    _escrever_linha_database(ws, 2, cotacoes, d_ref)
    wb.save(path); wb.close()
    log.info(f"DATABASE: nova linha para {d_ref}")


# ════════════════════════════════════════════════════════════════════════════
# 10. Main
# ════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Atualiza cotações SPOT CVU")
    parser.add_argument("--modo", choices=["pmo","revisao","ambos"], default="ambos")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--excel", type=str, default=None)
    args = parser.parse_args()

    global EXCEL_PATH
    if args.excel:
        EXCEL_PATH = Path(args.excel)

    eia_key = os.environ.get("EIA_API_KEY")
    jkm_key = os.environ.get("JKM_API_KEY")

    cotacoes = coletar_cotacoes(eia_key=eia_key, jkm_key=jkm_key)
    salvar_cache(cotacoes)
    imprimir_relatorio(cotacoes)

    if args.dry_run:
        log.info("Dry-run: Excel NÃO atualizado.")
    else:
        atualizar_excel(cotacoes, modo=args.modo)
        atualizar_database(cotacoes, EXCEL_PATH)
        log.info("=== Atualização concluída ===")


if __name__ == "__main__":
    main()
