"""
Câmbio BCB — duas abas via API PTAX do Banco Central do Brasil.

Fonte oficial (gratuita, sem chave):
  https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/

Reproduz o que a gestora vê nas páginas:
  - Histórico de cotações (Euro)  → aba "Cotações Moedas"
  - Fechamento diário do dólar    → aba "Fechamento Dólar"

Vantagem sobre scraping: é a MESMA fonte que as páginas do BCB usam, mas em
JSON estruturado. Sem browser, sem chave, sem defasagem (o boletim de
fechamento PTAX sai no mesmo dia útil).

ESCOPO DAS MOEDAS (CONFIRMAR COM A GESTORA):
  O CSV de referência era só Euro (código 978). MOEDAS define quais moedas
  entram na aba "Cotações Moedas". Para incluir mais, basta adicionar à lista
  (ex.: "GBP", "JPY", "CHF"). A API aceita qualquer símbolo do recurso Moedas.

Uso:
  python bcb_cambio.py --backfill    → histórico (padrão: ano corrente)
  python bcb_cambio.py               → atualiza com o dia mais recente
"""

import sys
import json
import time
import urllib.request
import urllib.parse
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "data/CME_Futures_Database.xlsx"

# Abas
SHEET_MOEDAS = "Cotações Moedas"
SHEET_DOLAR  = "Fechamento Dólar"

# Moedas da aba "Cotações Moedas" (CONFIRMAR ESCOPO COM A GESTORA).
# O arquivo de referência tinha apenas EUR.
MOEDAS = ["EUR"]

BASE = "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata"

# Formatação (mesmo padrão visual das outras abas)
HEADER_BG  = "2E4057"
HEADER_FG  = "FFFFFF"
ALT_ROW    = "EEF2F7"
BORDER_CLR = "B8C4CE"


# ---------------------------------------------------------------------------
# Chamadas à API PTAX
# ---------------------------------------------------------------------------

def _http_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
    })
    raw = urllib.request.urlopen(req, timeout=30).read()
    return json.loads(raw)


def _fmt_date(d: str) -> str:
    """'2026-07-14' -> 'MM-DD-YYYY' (formato exigido pela API PTAX)."""
    dt = datetime.strptime(d, "%Y-%m-%d")
    return dt.strftime("%m-%d-%Y")


def fetch_moeda_periodo(moeda: str, date_from: str, date_to: str) -> list:
    """
    Cotações de uma moeda num período. Retorna lista de dicts com o boletim
    de FECHAMENTO (tipoBoletim = 'Fechamento') de cada dia.
    """
    url = (f"{BASE}/CotacaoMoedaPeriodo("
           f"moeda=@moeda,dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
           f"?@moeda='{moeda}'"
           f"&@dataInicial='{_fmt_date(date_from)}'"
           f"&@dataFinalCotacao='{_fmt_date(date_to)}'"
           f"&$format=json")
    try:
        data = _http_json(url)
    except Exception as e:
        print(f"    [ERRO] {moeda}: {e}")
        return []

    out = []
    for v in data.get("value", []):
        # Mantém só o boletim de fechamento (evita os 5 boletins/dia)
        if v.get("tipoBoletim", "").strip().lower() != "fechamento":
            continue
        dt = str(v.get("dataHoraCotacao", ""))[:10]  # 'YYYY-MM-DD'
        if not dt:
            continue
        out.append({
            "data": dt,
            "moeda": moeda,
            "compra": v.get("cotacaoCompra"),
            "venda": v.get("cotacaoVenda"),
            "paridade_compra": v.get("paridadeCompra"),
            "paridade_venda": v.get("paridadeVenda"),
        })
    return out


def fetch_dolar_periodo(date_from: str, date_to: str) -> list:
    """Fechamento diário do dólar num período."""
    url = (f"{BASE}/CotacaoDolarPeriodo("
           f"dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
           f"?@dataInicial='{_fmt_date(date_from)}'"
           f"&@dataFinalCotacao='{_fmt_date(date_to)}'"
           f"&$format=json")
    try:
        data = _http_json(url)
    except Exception as e:
        print(f"    [ERRO] dólar: {e}")
        return []

    # NOTA: CotacaoDolarPeriodo NÃO retorna o campo tipoBoletim — cada data já
    # vem com um único registro (o fechamento). Não filtrar por tipoBoletim.
    out = []
    for v in data.get("value", []):
        dt = str(v.get("dataHoraCotacao", ""))[:10]
        if not dt:
            continue
        out.append({
            "data": dt,
            "compra": v.get("cotacaoCompra"),
            "venda": v.get("cotacaoVenda"),
        })
    return out


# ---------------------------------------------------------------------------
# Escrita nas abas
# ---------------------------------------------------------------------------

def _border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(ws, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "025C75"  # verde-petróleo do BCB


def _style_data_row(ws, row_idx, ncols):
    fill = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.fill = PatternFill("solid", start_color=fill)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()


def _num(v):
    """Converte para float aceitando string com vírgula ou ponto."""
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def update_moedas(rows: list):
    """Reconstrói a aba 'Cotações Moedas' (mais recente primeiro)."""
    wb = load_workbook(EXCEL_PATH)

    # merge com o que já existe (chave: data+moeda)
    existing = {}
    if SHEET_MOEDAS in wb.sheetnames:
        for r in wb[SHEET_MOEDAS].iter_rows(min_row=2, values_only=True):
            if r[0] and r[1]:
                existing[(str(r[0])[:10], str(r[1]))] = r
        del wb[SHEET_MOEDAS]

    merged = dict(existing)
    for r in rows:
        merged[(r["data"], r["moeda"])] = (
            r["data"], r["moeda"], _num(r["compra"]), _num(r["venda"]),
            _num(r["paridade_compra"]), _num(r["paridade_venda"]),
        )

    ws = wb.create_sheet(SHEET_MOEDAS)
    headers = ["Data", "Moeda", "Compra", "Venda", "Paridade Compra", "Paridade Venda"]
    _style_header(ws, headers, [14, 10, 14, 14, 16, 16])

    # ordena por data desc, depois moeda
    keys = sorted(merged.keys(), key=lambda k: (k[0], k[1]), reverse=True)
    for i, k in enumerate(keys, start=2):
        vals = merged[k]
        for col, val in enumerate(vals, 1):
            ws.cell(i, col, val)
        _style_data_row(ws, i, len(headers))

    wb.save(EXCEL_PATH)
    print(f"  [{SHEET_MOEDAS}] {len(rows)} linhas processadas ({len(merged)} no total)")


def update_dolar(rows: list):
    """Reconstrói a aba 'Fechamento Dólar' (mais recente primeiro)."""
    wb = load_workbook(EXCEL_PATH)

    existing = {}
    if SHEET_DOLAR in wb.sheetnames:
        for r in wb[SHEET_DOLAR].iter_rows(min_row=2, values_only=True):
            if r[0]:
                existing[str(r[0])[:10]] = r
        del wb[SHEET_DOLAR]

    merged = dict(existing)
    for r in rows:
        merged[r["data"]] = (r["data"], _num(r["compra"]), _num(r["venda"]))

    ws = wb.create_sheet(SHEET_DOLAR)
    headers = ["Data", "Compra", "Venda"]
    _style_header(ws, headers, [16, 16, 16])

    for i, d in enumerate(sorted(merged.keys(), reverse=True), start=2):
        for col, val in enumerate(merged[d], 1):
            ws.cell(i, col, val)
        _style_data_row(ws, i, len(headers))

    wb.save(EXCEL_PATH)
    print(f"  [{SHEET_DOLAR}] {len(rows)} linhas processadas ({len(merged)} no total)")


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def run(date_from: str = None, date_to: str = None):
    if date_to is None:
        date_to = date.today().isoformat()
    if date_from is None:
        date_from = (date.today() - timedelta(days=7)).isoformat()

    print(f"\n[Câmbio BCB] {date_from} → {date_to}")

    # Aba de moedas
    all_moedas = []
    for m in MOEDAS:
        print(f"  Moeda {m}...")
        rows = fetch_moeda_periodo(m, date_from, date_to)
        print(f"    {len(rows)} dias")
        all_moedas.extend(rows)
        time.sleep(0.5)
    if all_moedas:
        update_moedas(all_moedas)

    # Aba do dólar
    print("  Dólar (fechamento)...")
    dolar = fetch_dolar_periodo(date_from, date_to)
    print(f"    {len(dolar)} dias")
    if dolar:
        update_dolar(dolar)


if __name__ == "__main__":
    if "--backfill" in sys.argv:
        run(date_from=f"{date.today().year}-01-01")
    else:
        run()
